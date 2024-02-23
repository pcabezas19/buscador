from bs4 import BeautifulSoup
from selenium import webdriver
import time
import pandas as pd
import urllib.parse
import re
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

#***************EL MEJOR CON TODA LA FURIA *********************

# Función para realizar la extracción de datos
def scrape_data(root_urls):
    wb = Workbook()  # Crear un nuevo libro de trabajo de Excel
    
    while True:
        key_word = input("Ingresa nombre del producto para buscarlo o escribe 'salir' para finalizar: ")
        if key_word.lower() == 'salir':
            break  # Salir del bucle si el usuario ingresa 'no'
        
        # Crear una nueva hoja de cálculo para cada palabra clave
        ws = wb.create_sheet(title=key_word)
        
        # Codificar la palabra clave para que se ajuste a la URL
        encoded_key_word = urllib.parse.quote(key_word)
        
        # Diccionario para almacenar los datos por sitio web
        data_by_site = {}

        for site_name, root_url in root_urls.items():
            # Inicializar el navegador web
            driver = webdriver.Chrome(executable_path=r'C:\Users\thepa\Desktop\curso\Soifer2\chromedriver.exe')

            if site_name == 'tiendadirecta':
                parcial_url = f'{root_url}/search/?q={encoded_key_word}'
                class_name = 'item-description py-2 px-1'
                tag = 'div'
            elif site_name == 'thefoodmarket':
                parcial_url = f'{root_url}/{encoded_key_word}?_q={encoded_key_word}&map=ft'
                class_name = 'vtex-product-summary-2-x-container'
                tag = 'section'
            elif site_name == 'tiendanova':
                parcial_url = f'{root_url}/search/?q={encoded_key_word}'
                class_name = 'item-description py-4 px-1'
                tag = 'div'
            else:
                print(f"Nombre del sitio web no válido: {site_name}")
                continue

            # Cargar la página inicial
            driver.get(parcial_url)
            time.sleep(2)
            # Obtener el contenido HTML de la página después de hacer scroll down
            soup = BeautifulSoup(driver.page_source, 'html.parser')

            # Realizar scroll para cargar más elementos
            elementos = soup.find_all(tag, {'class': class_name})
            
            # Hacer scroll para cargar más elementos
            num_elementos_inicial = len(elementos)
            while True:
                # Hacer scroll lentamente hacia abajo
                for _ in range(10):
                    driver.execute_script("window.scrollBy(0, 100);")
                    time.sleep(0.07)  # Pausa de medio segundo entre cada scroll
            
                # Esperar un momento antes de capturar los nuevos elementos
                time.sleep(0.08)
            
                # Actualizar el contenido HTML después del scroll
                soup = BeautifulSoup(driver.page_source, 'html.parser')
                nuevos_elementos = soup.find_all(tag, {'class': class_name})
            
                # Hacer scroll lentamente hacia arriba
                for _ in range(5):
                    driver.execute_script("window.scrollBy(0, -50);")
                    time.sleep(0.2)  
            
                # Hacer scroll lentamente hacia abajo nuevamente
                for _ in range(10):
                    driver.execute_script("window.scrollBy(0, 100);")
            
                # Actualizar el contenido HTML después del segundo scroll hacia abajo
                soup = BeautifulSoup(driver.page_source, 'html.parser')
                nuevos_elementos = soup.find_all(tag, {'class': class_name})

                # Verificar si se han cargado nuevos elementos
                if len(nuevos_elementos) > num_elementos_inicial:
                    elementos = nuevos_elementos
                    num_elementos_inicial = len(elementos)
                else:
                    break

            # Extraer datos de los productos
            product_name = []
            product_price = []
            product_link = []
            product_promo = []
            for element in elementos:
                # Nombre del producto
                try:
                    if site_name == 'tiendadirecta':
                        name = element.find('div', {'class': 'js-item-name item-name mb-1'}).text.strip()
                        product_name.append(name)
                    elif site_name == 'thefoodmarket':
                        name = element.find('span').get_text().strip()
                        product_name.append(name)
                    elif site_name == 'tiendanova':
                        name = element.find('div', {'class': 'js-item-name item-name mb-3'}).text.strip()
                        product_name.append(name)
                    
                except AttributeError:
                    pass

                # Precio del producto
                try:
                    if site_name == 'tiendadirecta':
                        price = element.find('span', {'class': 'js-price-display item-price'}).text.strip()
                        
                    elif site_name == 'thefoodmarket':
                        price = element.find('span', {'class': 'vtex-product-price-1-x-currencyContainer'}).text.strip()
                        
                    elif site_name == 'tiendanova':
                        price = element.find('span', {'class': 'js-price-display item-price'}).text.strip()
                        
                    # Utilizar expresión regular para buscar el símbolo "$"
                    match = re.search(r'(?<=\$)\s*([\d,.]+)', price)  # Encontrar todos los dígitos, comas y puntos decimales después del símbolo "$" y opcionalmente espacios en blanco
                    if match:
                        price_value = match.group(1)  # Obtener el valor numérico del precio
                        # Establecer un formato específico para el precio
                        formatted_price = "${:,.2f}".format(float(price_value.replace('.', '').replace(',', '.')))  # Reemplazar las comas por puntos y formatear el precio
                        product_price.append(formatted_price)
                    else:
                        product_price.append(price)  # Si no se encuentra el símbolo "$", agregar el precio sin cambios
                except AttributeError:
                    pass
                
                # Promoción del producto
                try:
                    if site_name == 'tiendadirecta':
                        promo = element.find('div', {'class': 'js-offer-label label label-primary label-circle  small'}).text.strip()
                        product_promo.append(promo)
                    elif site_name == 'thefoodmarket':
                        promo= element.find('span', {'class': 'vtex-product-highlights-2-x-productHighlightText vtex-product-highlights-2-x-productHighlightText--highlights'}).text.strip()
                        product_promo.append(promo)
                    elif site_name == 'tiendanova':
                        promo = element.find('div', {'class': 'label-accent'}).text.strip()
                        product_promo.append(promo)
                    
                except AttributeError:
                    pass

                # Enlace del producto
                try:
                    if site_name == 'thefoodmarket':
                        dinamic_url = element.find('a', {'class': 'vtex-product-summary-2-x-clearLink'}).get('href')
                    elif site_name == 'tiendadirecta':
                        dinamic_url = element.find('a', {'class': 'item-link'}).get('href')
                    elif site_name == 'tiendanova':
                        dinamic_url = element.find('a', {'class': 'item-link'}).get('href')
                    product_link.append(urllib.parse.urljoin(root_url, dinamic_url).strip())
                    
                except AttributeError:
                    pass

            # Cerrar el navegador
            driver.quit()

            # Almacenar los datos en el diccionario por sitio web
            data_by_site[site_name] = {
                'Name': product_name,
                'Price': product_price,
                'Promo': product_promo,
                'Link': product_link
            }

        # Obtener longitudes máximas de todas las listas
        max_lengths = {site_name: max(len(data['Name']), len(data['Price']), len(data['Promo']), len(data['Link'])) for site_name, data in data_by_site.items()}

        # Rellenar las listas con valores vacíos para que tengan la misma longitud
        for site_name, data in data_by_site.items():
            data['Name'] += [''] * (max_lengths[site_name] - len(data['Name']))
            data['Price'] += [''] * (max_lengths[site_name] - len(data['Price']))
            data['Promo'] += [''] * (max_lengths[site_name] - len(data['Promo']))
            data['Link'] += [''] * (max_lengths[site_name] - len(data['Link']))

        # Crear el DataFrame combinado
        combined_data = pd.concat([pd.DataFrame(data) for data in data_by_site.values()], axis=0)
        
        # Ordenar el DataFrame por precio de mayor a menor
        combined_data = combined_data.sort_values(by='Price', ascending=True)

        # Establecer los títulos en la primera fila
        titles = ['Name', 'Price', 'Promo', 'Link']
        for col, title in enumerate(titles, start=1):
            ws.cell(row=1, column=col, value=title)
        
        # Llenar la hoja de cálculo con los datos
        for r_idx, row in enumerate(combined_data.iterrows(), 2):
            for c_idx, value in enumerate(row[1], 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        # Ajustar automáticamente el ancho de las columnas
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2  # Añadir un poco de espacio adicional
    # Eliminar la hoja de Excel en blanco
    wb.remove(wb['Sheet'])
    # Guardar el archivo Excel
    wb.save(r'C:\Users\thepa\Desktop\curso\Soifer2\comparativaJessy.xlsx')

    print("Los resultados se han guardado en el archivo 'resultados_multi_page.xlsx'.")

# Ejemplo de uso
root_urls = {
    'thefoodmarket': 'https://www.thefoodmarket.com.ar',
    'tiendadirecta': 'https://www.tiendadirecta.com.ar',
    'tiendanova': 'https://www.tiendanova.com'
}

scrape_data(root_urls)
